import pandas as pd
import ezodf
import pyexcel
import os
import xlsxwriter
import openpyxl

def read_create():
    file_name = "src/data/Courselist.ods"

    sheet =pyexcel.get_sheet(file_name=file_name)
    i=0
    for row in sheet:
    
        lists=str(row[3])
        lists=lists.split('-')
        if(lists[0]!="0" ):
            for p in range(len(row)):
                if(row[p]=='C'):    
                    if(os.path.exists("src/tmp/baskets/"+str(p) + ".xlsx")  ):
                        wb = openpyxl.load_workbook("src/tmp/baskets/"+str(p) + ".xlsx")
                        pass
                    else:
                        wb = openpyxl.Workbook()
                        wb.save("src/tmp/baskets/"+str(p) + ".xlsx")
                        wb = openpyxl.load_workbook("src/tmp/baskets/"+str(p) + ".xlsx")
                    ws = wb.worksheets[0]
                    rows = ws.max_row+1
                    for l in range(1,6):
                        ws.cell(rows,l).value=row[l]
                        wb.save("src/tmp/baskets/"+str(p) + ".xlsx")
                    ws.cell(rows,6).value=row[p]
                    wb.save("src/tmp/baskets/"+str(p) + ".xlsx")

read_create()